import pandas as pd
import openpyxl
from openpyxl.styles import Font
import re

def create_df(table):
                        rows = table.find_all('tr')
                        data = []
                        col_spans = []
                        row_spans = []
                        # to find the rowspan and colspan
                        for i, row in enumerate(rows):
                            cells = row.find_all(['th', 'td'])
                            row_data = []
                            for j, cell in enumerate(cells):
                                if cell.has_attr('colspan'):
                                    col_spans.append((i, j, int(cell['colspan'])))
                                if cell.has_attr('rowspan'):
                                    row_spans.append((i, j, int(cell['rowspan'])))
                                row_data.append(cell.get_text(strip=True))
                            data.append(row_data)
                        # Apply colspans to data
                        for span in col_spans:
                            i, j, span_size = span
                            for r in range(1, span_size):
                                data[i].insert(j+1, '')
                        # Apply rowspans to data
                        for span in row_spans:
                            i, j, span_size = span
                            for r in range(1, span_size):
                                data[i+r].insert(j, '')

                        # Convert the data into a pandas dataframe
                        df = pd.DataFrame(data[1:], columns=data[0])

                        # Replace any None values with an empty string
                        df = df.fillna('')
                        # convert the dataframe to dict
                        data_dict = {}
                        # convert the dataframe to dict
                        data_dict = df.to_dict('list')
                        #print(data_dict)
                        return(data_dict)


def tc_details (h1_tags,sheet1,a):
    if a == 0: 
        head = ["S.no" ,"Test Plan","Cluster Name",	"TC ID","TC Name","TC Full Name"]
        sheet1.append(head)
        next_row = sheet1.max_row + 1
        appended_row = sheet1[next_row]

    # Make the entire row bold
        for cell in appended_row:
            cell.font = Font(bold=True)
    next_row = sheet1.max_row + 1
    s=0
    for i in range(len(h1_tags)):
        
        h1 = h1_tags[i].text
        cluster_name = h1.replace('Cluster Test Plan', '')

        first_h1 = h1_tags[i]  # Select the first h1 tag in the pair
        if i == (len(h1_tags)-1):
            second_h1 = False
        else:
            second_h1 = h1_tags[i+1]  # Select the second h1 tag in the pair
            
        # Find all h5 tags after the first h1 tag with id starting with the value
        h5_tags = first_h1.find_all_next('h5', {'id': lambda x: x and x.startswith('_test_procedure')})  
        result = []
        if second_h1:
            for h5_tag in h5_tags:
                #print(h5_tag)
                if h5_tag.find_previous('h1') == first_h1:
                    if h5_tag.find_next('h1') == second_h1:
                        result.append(h5_tag)
        
        else:
            for h5_tag in h5_tags:
                result.append(h5_tag)

        if result:
               
                heads = []
                for h5_tag in result:
                    h4_tag = h5_tag.find_previous('h4')
                    headt= h4_tag.text
                    heads.append(headt)
                    index = result.index(h5_tag)
                    #print (index)

                    
        else:

                h5_tags = first_h1.find_all_next('h6', {'id': lambda x: x and x.startswith('_test_procedure')})  

                for h5_tag in h5_tags:
                #print(h5_tag)
                    if h5_tag.find_previous('h1') == first_h1:
                        if h5_tag.find_next('h1') == second_h1:
                            result.append(h5_tag)
                
                print (result)
                heads =[]
                for h5_tag in result:
                    h4_tag = h5_tag.find_previous('h5')
                    headt= h4_tag.text
                    heads.append(headt)
                    index = result.index(h5_tag)
                    #print (index)


        
        for j in range(len(result)):
            testcase1 = re.search(r'\[(.*?)\]',heads[j])
            if testcase1:
                        matched_str = testcase1.group()  # Extract the matched substring
                        testcase = re.sub(r'\[|\]', '', matched_str)

            tc_name = heads[j].split("]", 1)[-1].strip()

            if a == 0:
                tp = "Main-Test-Plan"
            else:
                tp  = "App-Cluster" 
            
            if i == 0 & j ==0 :
                n = 1
            else:
                n = sheet1.max_row - 1
            

            value = [n , tp , cluster_name, testcase,tc_name, heads[j] ]
            print(value)
            

            sheet1.append(value)
        
