from bs4 import BeautifulSoup
from script.makehtml import file
from script.diff import *
from script.create_data import dict_
from script.createtable import *
import pandas as pd
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import Alignment
import datetime

workbook = openpyxl.Workbook()
today = datetime.date.today().strftime('%Y-%m-%d')
filename = f"test_plan_change_{today}.xlsx"


def test(h1_tags):
      for i in range(len(h1_tags)-1):
            h1 = h1_tags[i].text
            sheet=h1.replace('Cluster Test Plan', '').replace(' ', '').replace('/', '')
            print(sheet)
            first_h1 = h1_tags[i]  # Select the first h1 tag in the pair
            second_h1 = h1_tags[i+1]  # Select the second h1 tag in the pair
            
            # Find all h5 tags after the first h1 tag with id starting with the value
            h5_tags = first_h1.find_all_next('h5', {'id': lambda x: x and x.startswith('_test_procedure')})  
            
            
            new_sheet = workbook.create_sheet(sheet)
            result = []

            # Find the h4 tag of the paeticular cluster
            for h5_tag in h5_tags:
                #print(h5_tag)
                if h5_tag.find_previous('h1') == first_h1:
                    if h5_tag.find_next('h1') == second_h1:
                        result.append(h5_tag)
            
            print (result)

            
            data_tables = []
            
            if result:
               
                heads = []
                for h5_tag in result:
                    h4_tag = h5_tag.find_previous('h4')
                    headt= h4_tag.text
                    heads.append(headt)
                                   

                    table = h5_tag.find_next('table')
                    if table:

                        data_dict = create_df(table)
                        data_tables.append(data_dict)
                        
                    else: 
                        print("no table no table no table")

                

                    
            else:

                h5_tags = first_h1.find_all_next('h6', {'id': lambda x: x and x.startswith('_test_procedure')})  

                for h5_tag in h5_tags:
                #print(h5_tag)
                    if h5_tag.find_previous('h1') == first_h1:
                        if h5_tag.find_next('h1') == second_h1:
                            result.append(h5_tag)
                
                print (result)
                heads =[]
                for h5_tag in result:
                    h4_tag = h5_tag.find_previous('h5')
                    headt= h4_tag.text
                    heads.append(headt)
                    

                    table = h5_tag.find_next('table')

                    print(table)
                    if table:
                        data_dict = create_df(table)
                        data_tables.append(data_dict)
                    else:
                        print("no table no table no table")
                    
            print (data_tables)   
                    

                    
            
                                
            # list to store the col of testcase
            if data_tables:
                            for i in range(len(heads)):
                                    data_table = data_tables[i]
                                    head = heads[i]

                                    print (data_table)
                                    list_a = data_table.get('#',data_table.get('Item'))
                                    list_b = data_table['Ref']
                                    list_c = data_table['PICS']
                                    list_d = data_table.get('Test Step',data_table.get('Test Harness Step'))
                                    list_e = data_table.get('Expected Outcome', data_table.get('DUT Pass Verification'))
                                            
                                    print(list_a)
                                    print(list_b)
                                    length = len(list_a)

                                            
                                    
                                    new_sheet.append([""])

                                    new_sheet.append([head])
                                    new_sheet.append([""])
                                    keys_list = list(data_dict.keys())

                                    new_sheet.append(keys_list)
                                    
                                    # writing the headers of testcase
                                    
                                    # writing the content of testcase
                                    for i in range(len(list_a)):
                                        val =[list_a[i], list_b[i],list_c[i], list_d[i],list_e[i]]
                                        new_sheet.append(val)



                                    column_widths = {'A': 10, 'B': 20, 'C': 20 ,'D':30,'E':30}  # Specify the column widths as desired

                                    for column, width in column_widths.items():
                                                new_sheet.column_dimensions[column].width = width
                                        
                                    workbook.save(filename) 
                        
                    

                        
            else:
                        print("no table found")


if __name__ == '__main__':
   
    print("ready")
    html_file = file()
    workbook.save(filename)
    app =  html_file['app']
    main = html_file['main']


    sheet1 = workbook.active
    sheet1.title = "All_TC_Details"
    sheet2 = workbook.create_sheet("changes_in_cluster")
    sheet3 = workbook.create_sheet("changes_in_tc")

    for a in range(len(app)):
        if a == 0:
            files = app
        else:
            files = main

        html1 = files[0]
        html2 = files[1]

        with open (html1) as f:
                soup1 = BeautifulSoup(f, 'html.parser')
            
        with open (html2) as f:
                soup2 = BeautifulSoup(f, 'html.parser')


        #list of clusters
        h1_tags1 = soup1.find_all('h1', {'id': True}) 
        h1_tags2 = soup2.find_all('h1', {'id': True})
        h1text1 = []
        h1text2 = []

        for h1_tag in h1_tags1:
                h1_text1= h1_tag.text
                print (h1_text1)
                h1text1.append(h1_text1)

        for h1_tag in h1_tags2:
                h1_text2= h1_tag.text
                print (h1_text2)
                h1text2.append(h1_text2)
        
        

        tc_details(h1_tags1,sheet1,a)
        column_widths = {'A': 10, 'B': 20, 'C': 20 ,'D':30,'E':40,"F":50}  # Specify the column widths as desired

        for column, width in column_widths.items():
                sheet1.column_dimensions[column].width = width
        workbook.save(filename)

        new_data_dicts = dict_(h1_tags1)
        old_data_dicts = dict_(h1_tags2)

        new_data_dicts_keys = list(new_data_dicts.keys())
        old_data_dicts_keys = list(old_data_dicts.keys())


        cluster_dif(h1text1,h1text2,new_data_dicts_keys,old_data_dicts_keys,sheet2)
        workbook.save(filename)
        

        added_tc, removed_tc, mod_tc = tc_diff(new_data_dicts_keys, old_data_dicts_keys)
        data = [new_data_dicts_keys, old_data_dicts_keys, new_data_dicts, old_data_dicts, removed_tc,added_tc,mod_tc, sheet3]

        dif_table(data)
        test(h1_tags1)
        

        for sheet in workbook:
        # Iterate through each row in the sheet
                for row in sheet.iter_rows():
                # Iterate through each cell in the row
                        for cell in row:
                        # Apply text wrap formatting to the cell
                                cell.alignment = Alignment(wrapText=True)
        

        workbook.save(filename)





